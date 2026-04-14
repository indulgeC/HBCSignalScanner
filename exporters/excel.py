"""
Exporter — writes signals to Excel (.xlsx) and CSV.
"""

from __future__ import annotations
import csv
import logging
import os
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models.signal import Signal

logger = logging.getLogger(__name__)

# ── Column config ─────────────────────────────────────────────────────────────

DISPLAY_HEADERS = [
    ("Signal Title", 40),
    ("Agency", 25),
    ("Geography", 20),
    ("Sector", 14),
    ("Estimated Value", 18),
    ("Expected Timeline", 18),
    ("Meeting Date", 16),
    ("Signal Type", 22),
    ("Procurement Stage", 28),
    ("Lifecycle Stage", 24),
    ("Signal Strength", 16),
    ("Strategic Fit", 16),
    ("Friction Level", 14),
    ("Momentum", 14),
    ("Trigger Event", 40),
    ("Strategic Notes", 55),
    ("Source Link", 50),
]

AUDIT_HEADERS = [
    ("Evidence Snippet", 50),
    ("Evidence Page", 12),
    ("Confidence", 10),
    ("Method", 10),
    ("File URL", 40),
    ("Page URL", 40),
]

FIELD_MAP = [
    "signal_title", "agency", "geography", "sector",
    "estimated_value", "expected_timeline", "meeting_date",
    "signal_type", "procurement_stage", "lifecycle_stage",
    "signal_strength", "strategic_fit", "friction_level", "momentum",
    "trigger_event", "strategic_notes", "source_link",
]

AUDIT_FIELD_MAP = [
    "evidence_snippet", "evidence_page", "confidence_score",
    "extraction_method", "source_file_url", "source_page_url",
]


# ── Colors ────────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
AUDIT_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

STRENGTH_COLORS = {
    "High": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "Medium": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "Low": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}

FIT_COLORS = {
    "Strong Fit": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "Moderate Fit": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "Monitor": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
    "No Fit": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}

MOMENTUM_COLORS = {
    "Accelerating": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "Stable": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
    "Stalled": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "Unclear": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
}

FRICTION_COLORS = {
    "Low": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "Moderate": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "High": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


# ── Excel export ──────────────────────────────────────────────────────────────

def export_excel(
    signals: List[Signal],
    output_path: str,
    include_audit: bool = True,
) -> str:
    """Write signals to a formatted .xlsx file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Signals"

    all_headers = list(DISPLAY_HEADERS)
    all_fields = list(FIELD_MAP)
    if include_audit:
        all_headers += AUDIT_HEADERS
        all_fields += AUDIT_FIELD_MAP

    # ── Header row ────────────────────────────────────────────────────
    for col_idx, (header, width) in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        fill = HEADER_FILL if col_idx <= len(DISPLAY_HEADERS) else AUDIT_HEADER_FILL
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(all_headers))}1"

    # ── Data rows ─────────────────────────────────────────────────────
    body_font = Font(name="Calibri", size=10)
    wrap_align = Alignment(vertical="top", wrap_text=True)

    for row_idx, sig in enumerate(signals, 2):
        for col_idx, field_name in enumerate(all_fields, 1):
            value = getattr(sig, field_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.alignment = wrap_align
            cell.border = THIN_BORDER

            # Conditional formatting
            if field_name == "signal_strength" and value in STRENGTH_COLORS:
                cell.fill = STRENGTH_COLORS[value]
            elif field_name == "strategic_fit" and value in FIT_COLORS:
                cell.fill = FIT_COLORS[value]
            elif field_name == "momentum" and value in MOMENTUM_COLORS:
                cell.fill = MOMENTUM_COLORS[value]
            elif field_name == "friction_level" and value in FRICTION_COLORS:
                cell.fill = FRICTION_COLORS[value]
            elif field_name == "source_link" and value:
                cell.hyperlink = value
                cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    logger.info("Excel saved: %s  (%d signals)", output_path, len(signals))
    return output_path


# ── CSV export ────────────────────────────────────────────────────────────────

def export_csv(signals: List[Signal], output_path: str) -> str:
    """Write signals to a CSV file."""
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    headers = [h for h, _ in DISPLAY_HEADERS]

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for sig in signals:
            row = [getattr(sig, field, "") for field in FIELD_MAP]
            writer.writerow(row)

    logger.info("CSV saved: %s  (%d signals)", output_path, len(signals))
    return output_path
